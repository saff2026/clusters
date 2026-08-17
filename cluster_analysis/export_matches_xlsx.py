#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""يصدّر جداول المباريات (المنطقة ← الفئة ← المجموعة) إلى ملف إكسل.
نفس منطق صفحة matches.html: المجموعات المكتملة فقط (الفرق ≥ الهدف)،
والفرق والمباريات مأخوذة من صفحة «عدد المباريات» في الإكسل (matches_data.json).
الاستخدام: python3 export_matches_xlsx.py [مسار_الإخراج.xlsx]
"""
import sys, json
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = "/home/user/khitba/cluster_analysis/"
OUT = sys.argv[1] if len(sys.argv) > 1 else BASE + "جداول_المباريات.xlsx"

T = json.load(open(BASE + "teams2.json", encoding="utf-8"))
M = json.load(open(BASE + "_maps.json", encoding="utf-8"))
STRUCT = M["STRUCT"]
TARGET = T.get("target", 6)
MDATA = json.load(open(BASE + "matches_data.json", encoding="utf-8"))

# مجموعة -> مدنها ومنطقتها
gc = defaultdict(set)
greg = {}
for rg in STRUCT:
    for ag in STRUCT[rg]:
        for g, cities in STRUCT[rg][ag].items():
            for c in cities:
                gc[g].add(c)
            greg.setdefault(g, rg)

def teams_of(age, g):
    d = MDATA.get(age + "|" + g)
    return d["n"] if d else 0
def matches_of(age, g):
    d = MDATA.get(age + "|" + g)
    return d["m"] if d else 0

groups_by_age = defaultdict(list)
for key in MDATA:
    a, g = key.split("|", 1)
    groups_by_age[a].append(g)

# المجموعات المكتملة لكل فئة
byAge = {}
for age in T["ages"]:
    arr = []
    for g in groups_by_age.get(age, []):
        n = teams_of(age, g)
        if n >= TARGET and g != "(غير مصنّف)":
            arr.append({"group": g, "region": greg.get(g, "غير محدد"),
                        "cities": "، ".join(sorted(gc.get(g, []))),
                        "n": n, "matches": matches_of(age, g)})
    arr.sort(key=lambda x: -x["matches"])
    byAge[age] = arr

# تجميع المنطقة ← الفئة ← المجموعة
R = {}
for age in T["ages"]:
    for x in byAge[age]:
        R.setdefault(x["region"], {}).setdefault(age, []).append(x)

# ===== كتابة الإكسل =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "جداول المباريات"
ws.sheet_view.rightToLeft = True

GREEN = "0B3D2E"
GOLD = "FFD166"
LGREEN = "DDF3E7"
thin = Side(style="thin", color="BFDFCF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
right = Alignment(horizontal="right", vertical="center")

headers = ["المنطقة", "الفئة", "المجموعة", "المدن", "عدد الفرق", "عدد المباريات"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill("solid", fgColor=GREEN)
    cell.alignment = center
    cell.border = border

grand = 0
grand_teams = 0
for rg in sorted(R.keys()):
    reg_tot = sum(x["matches"] for age in T["ages"] for x in R[rg].get(age, []))
    reg_teams = sum(x["n"] for age in T["ages"] for x in R[rg].get(age, []))
    grand += reg_tot
    grand_teams += reg_teams
    for age in T["ages"]:
        for x in R[rg].get(age, []):
            ws.append([rg, age, x["group"], x["cities"], x["n"], x["matches"]])
            r = ws.max_row
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = center if c >= 5 else right
            ws.cell(row=r, column=3).font = Font(bold=True)
            ws.cell(row=r, column=6).font = Font(bold=True, color="9A6B00")
    # سطر إجمالي المنطقة
    ws.append([rg + " — الإجمالي", "", "", "", reg_teams, reg_tot])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    for c in range(1, 7):
        cell = ws.cell(row=r, column=c)
        cell.fill = PatternFill("solid", fgColor=LGREEN)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = center if c >= 5 else right

# سطر الإجمالي الكلي
ws.append(["الإجمالي الكلي", "", "", "", grand_teams, grand])
r = ws.max_row
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
for c in range(1, 7):
    cell = ws.cell(row=r, column=c)
    cell.fill = PatternFill("solid", fgColor=GOLD)
    cell.font = Font(bold=True, size=12)
    cell.border = border
    cell.alignment = center if c >= 5 else right

widths = [20, 10, 26, 55, 12, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"

wb.save(OUT)
print("تم حفظ:", OUT)
print("مجموع المباريات:", grand, "| مجموع الفرق:", grand_teams,
      "| مناطق:", len(R))
